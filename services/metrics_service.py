import hashlib
import os
import re
import sqlite3
import uuid
from difflib import SequenceMatcher
from datetime import datetime
from typing import Any, Dict, List, Optional

from openpyxl import Workbook


RUN_COLUMNS = [
    "run_id",
    "timestamp",
    "problem_number",
    "problem_name",
    "problem_link",
    "difficulty",
    "language",
    "model",
    "prompt_version",
    "prompt_strategy",
    "prompt_hash",
    "prompt_preview",
    "prompt_text",
    "prompt_chars",
    "prompt_lines",
    "code_chars",
    "code_lines",
    "code_sha256",
    "code_text",
    "prompt_tokens",
    "response_tokens",
    "total_tokens",
    "output_input_ratio",
    "total_duration_ms",
    "load_duration_ms",
    "prompt_eval_ms",
    "generation_ms",
    "tokens_per_sec",
    "http_status",
    "error_type",
    "retry_count",
    "timeout_flag",
    "llm_returned_code_block",
    "code_appended_externally",
    "llm_response_chars",
    "llm_response_lines",
    "llm_response_text",
    "response_chars",
    "response_lines",
    "has_title",
    "has_intuition",
    "has_approach",
    "has_time_complexity",
    "has_space_complexity",
    "has_code_block",
    "format_score",
    "completeness_score",
    "manual_edit_distance",
    "accepted_for_posting",
    "error_message",
]


MIGRATION_COLUMNS = {
    "problem_link": "TEXT",
    "prompt_strategy": "TEXT",
    "prompt_hash": "TEXT",
    "prompt_preview": "TEXT",
    "prompt_text": "TEXT",
    "code_sha256": "TEXT",
    "code_text": "TEXT",
    "llm_returned_code_block": "INTEGER",
    "code_appended_externally": "INTEGER",
    "llm_response_chars": "INTEGER",
    "llm_response_lines": "INTEGER",
    "llm_response_text": "TEXT",
}


def _base_dir() -> str:
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def get_metrics_paths() -> Dict[str, str]:
    stats_dir = os.path.join(_base_dir(), "llm_stats")
    return {
        "stats_dir": stats_dir,
        "db_path": os.path.join(stats_dir, "runs.db"),
        "excel_path": os.path.join(stats_dir, "token_usage.xlsx"),
    }


def _connect() -> sqlite3.Connection:
    paths = get_metrics_paths()
    os.makedirs(paths["stats_dir"], exist_ok=True)
    conn = sqlite3.connect(paths["db_path"])
    conn.row_factory = sqlite3.Row
    return conn


def ensure_metrics_storage() -> None:
    conn = _connect()
    try:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS llm_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                run_id TEXT UNIQUE NOT NULL,
                timestamp TEXT NOT NULL,
                problem_number TEXT,
                problem_name TEXT,
                problem_link TEXT,
                difficulty TEXT,
                language TEXT,
                model TEXT,
                prompt_version TEXT,
                prompt_strategy TEXT,
                prompt_hash TEXT,
                prompt_preview TEXT,
                prompt_text TEXT,
                prompt_chars INTEGER,
                prompt_lines INTEGER,
                code_chars INTEGER,
                code_lines INTEGER,
                code_sha256 TEXT,
                code_text TEXT,
                prompt_tokens INTEGER,
                response_tokens INTEGER,
                total_tokens INTEGER,
                output_input_ratio REAL,
                total_duration_ms REAL,
                load_duration_ms REAL,
                prompt_eval_ms REAL,
                generation_ms REAL,
                tokens_per_sec REAL,
                http_status INTEGER,
                error_type TEXT,
                retry_count INTEGER,
                timeout_flag INTEGER,
                llm_returned_code_block INTEGER,
                code_appended_externally INTEGER,
                llm_response_chars INTEGER,
                llm_response_lines INTEGER,
                llm_response_text TEXT,
                response_chars INTEGER,
                response_lines INTEGER,
                has_title INTEGER,
                has_intuition INTEGER,
                has_approach INTEGER,
                has_time_complexity INTEGER,
                has_space_complexity INTEGER,
                has_code_block INTEGER,
                format_score REAL,
                completeness_score REAL,
                manual_edit_distance INTEGER,
                accepted_for_posting INTEGER,
                error_message TEXT
            )
            """
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_llm_runs_timestamp ON llm_runs(timestamp)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_llm_runs_problem ON llm_runs(problem_number)"
        )
        _migrate_metrics_schema(conn)
        conn.commit()
    finally:
        conn.close()


def _migrate_metrics_schema(conn: sqlite3.Connection) -> None:
    existing = {row[1] for row in conn.execute("PRAGMA table_info(llm_runs)").fetchall()}
    for column_name, column_type in MIGRATION_COLUMNS.items():
        if column_name not in existing:
            conn.execute(f"ALTER TABLE llm_runs ADD COLUMN {column_name} {column_type}")


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        if value is None:
            return default
        return int(value)
    except (TypeError, ValueError):
        return default


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value is None:
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


def _ns_to_ms(nanos: Any) -> float:
    return _safe_float(nanos) / 1_000_000


def _sha256_hex(value: str) -> str:
    return hashlib.sha256((value or "").encode("utf-8")).hexdigest()


def _build_prompt_preview(prompt: str, max_len: int = 280) -> str:
    compact = " ".join((prompt or "").split())
    if len(compact) <= max_len:
        return compact
    return compact[: max_len - 3] + "..."


# Max column width cap — long text fields (prompt_text, code_text, etc.) can be
# thousands of chars; we cap at 80 so the sheet stays usable.
_EXCEL_MAX_COL_WIDTH = 80
_EXCEL_MIN_COL_WIDTH = 10


def _autofit_columns(ws) -> None:
    """Set each column width to fit its header + a small margin, capped at 80."""
    for col_cells in ws.iter_cols():
        header = str(col_cells[0].value or "")
        # Sample up to 20 data rows to find a representative value width.
        sample_max = max(
            (len(str(cell.value or "")) for cell in col_cells[1:21]),
            default=0,
        )
        width = max(len(header) + 4, min(sample_max + 2, _EXCEL_MAX_COL_WIDTH), _EXCEL_MIN_COL_WIDTH)
        ws.column_dimensions[col_cells[0].column_letter].width = width


def _prepare_export_row(row: Dict[str, Any]) -> Dict[str, Any]:
    """Normalize missing legacy fields so Excel output is analysis-friendly."""
    normalized = dict(row)

    _LEGACY_MARKER = "legacy"
    _DASH = "—"

    if not normalized.get("prompt_strategy"):
        normalized["prompt_strategy"] = _LEGACY_MARKER

    if not normalized.get("problem_link"):
        normalized["problem_link"] = _DASH

    if not normalized.get("prompt_text"):
        normalized["prompt_text"] = _DASH

    if not normalized.get("code_text"):
        normalized["code_text"] = _DASH

    if not normalized.get("llm_response_text"):
        normalized["llm_response_text"] = _DASH

    if not normalized.get("prompt_preview"):
        raw = normalized.get("prompt_text", "")
        normalized["prompt_preview"] = _build_prompt_preview(raw) if raw != _DASH else _DASH

    if not normalized.get("prompt_hash"):
        raw = normalized.get("prompt_text", "")
        normalized["prompt_hash"] = _sha256_hex(raw) if raw != _DASH else _DASH

    if not normalized.get("code_sha256"):
        raw = normalized.get("code_text", "")
        normalized["code_sha256"] = _sha256_hex(raw) if raw != _DASH else _DASH

    if not normalized.get("llm_response_chars"):
        raw = normalized.get("llm_response_text", "")
        normalized["llm_response_chars"] = len(raw) if raw != _DASH else 0

    if not normalized.get("llm_response_lines"):
        raw = normalized.get("llm_response_text", "")
        normalized["llm_response_lines"] = len(raw.splitlines()) if raw != _DASH else 0

    return normalized


def analyze_response_quality(text: str) -> Dict[str, Any]:
    content = (text or "").strip()
    first_non_empty = ""
    for line in content.splitlines():
        if line.strip():
            first_non_empty = line.strip()
            break

    has_title = bool(first_non_empty and not first_non_empty.startswith("#"))
    has_intuition = bool(re.search(r"^##\s+Intuition\b", content, flags=re.IGNORECASE | re.MULTILINE))
    has_approach = bool(re.search(r"^##\s+Approach\b", content, flags=re.IGNORECASE | re.MULTILINE))
    has_time_complexity = bool(
        re.search(r"^##\s+Time\s+Complexity\b", content, flags=re.IGNORECASE | re.MULTILINE)
    )
    has_space_complexity = bool(
        re.search(r"^##\s+Space\s+Complexity\b", content, flags=re.IGNORECASE | re.MULTILINE)
    )
    has_code_block = "```" in content

    checks = [
        has_title,
        has_intuition,
        has_approach,
        has_time_complexity,
        has_space_complexity,
        has_code_block,
    ]
    completeness_score = round((sum(int(x) for x in checks) / len(checks)) * 100, 2)

    # Weighted score favors correctly formatted markdown structure.
    format_score = round(
        (20 * int(has_title))
        + (15 * int(has_intuition))
        + (15 * int(has_approach))
        + (15 * int(has_time_complexity))
        + (15 * int(has_space_complexity))
        + (20 * int(has_code_block)),
        2,
    )

    return {
        "has_title": int(has_title),
        "has_intuition": int(has_intuition),
        "has_approach": int(has_approach),
        "has_time_complexity": int(has_time_complexity),
        "has_space_complexity": int(has_space_complexity),
        "has_code_block": int(has_code_block),
        "format_score": format_score,
        "completeness_score": completeness_score,
    }


def build_run_record(
    *,
    problem_number: str,
    problem_name: str,
    difficulty: str,
    language: str,
    model: str,
    prompt_version: str,
    prompt: str,
    code: str,
    response_text: str,
    problem_link: str = "",
    prompt_strategy: str = "",
    llm_response_text: Optional[str] = None,
    response_data: Optional[Dict[str, Any]] = None,
    http_status: Optional[int] = None,
    error_type: str = "",
    error_message: str = "",
    llm_returned_code_block: Optional[int] = None,
    code_appended_externally: int = 0,
    retry_count: int = 0,
    timeout_flag: int = 0,
    manual_edit_distance: Optional[int] = None,
    accepted_for_posting: Optional[int] = None,
) -> Dict[str, Any]:
    response_data = response_data or {}
    llm_text = llm_response_text if llm_response_text is not None else response_text
    final_output = response_text or ""

    prompt_tokens = _safe_int(response_data.get("prompt_eval_count"))
    response_tokens = _safe_int(response_data.get("eval_count"))
    total_tokens = prompt_tokens + response_tokens

    total_duration_ns = response_data.get("total_duration", 0)
    load_duration_ns = response_data.get("load_duration", 0)
    prompt_eval_ns = response_data.get("prompt_eval_duration", 0)
    generation_ns = response_data.get("eval_duration", 0)

    output_input_ratio = 0.0
    if prompt_tokens > 0:
        output_input_ratio = round(response_tokens / prompt_tokens, 4)

    tokens_per_sec = 0.0
    generation_seconds = _safe_float(generation_ns) / 1_000_000_000
    if generation_seconds > 0:
        tokens_per_sec = round(response_tokens / generation_seconds, 4)

    quality = analyze_response_quality(final_output)
    prompt_hash = _sha256_hex(prompt or "")
    code_sha256 = _sha256_hex(code or "")

    if llm_returned_code_block is None:
        llm_returned_code_block = int("```" in (llm_text or "") or "## Code" in (llm_text or ""))

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    return {
        "run_id": str(uuid.uuid4()),
        "timestamp": now,
        "problem_number": str(problem_number or ""),
        "problem_name": str(problem_name or ""),
        "problem_link": str(problem_link or ""),
        "difficulty": str(difficulty or ""),
        "language": str(language or ""),
        "model": str(model or ""),
        "prompt_version": str(prompt_version or ""),
        "prompt_strategy": str(prompt_strategy or ""),
        "prompt_hash": prompt_hash,
        "prompt_preview": _build_prompt_preview(prompt),
        "prompt_text": str(prompt or ""),
        "prompt_chars": len(prompt or ""),
        "prompt_lines": len((prompt or "").splitlines()),
        "code_chars": len(code or ""),
        "code_lines": len((code or "").splitlines()),
        "code_sha256": code_sha256,
        "code_text": str(code or ""),
        "prompt_tokens": prompt_tokens,
        "response_tokens": response_tokens,
        "total_tokens": total_tokens,
        "output_input_ratio": output_input_ratio,
        "total_duration_ms": round(_ns_to_ms(total_duration_ns), 2),
        "load_duration_ms": round(_ns_to_ms(load_duration_ns), 2),
        "prompt_eval_ms": round(_ns_to_ms(prompt_eval_ns), 2),
        "generation_ms": round(_ns_to_ms(generation_ns), 2),
        "tokens_per_sec": tokens_per_sec,
        "http_status": _safe_int(http_status, default=0),
        "error_type": str(error_type or ""),
        "retry_count": _safe_int(retry_count, default=0),
        "timeout_flag": _safe_int(timeout_flag, default=0),
        "llm_returned_code_block": _safe_int(llm_returned_code_block, default=0),
        "code_appended_externally": _safe_int(code_appended_externally, default=0),
        "llm_response_chars": len(llm_text or ""),
        "llm_response_lines": len((llm_text or "").splitlines()),
        "llm_response_text": str(llm_text or ""),
        "response_chars": len(final_output),
        "response_lines": len(final_output.splitlines()),
        "has_title": quality["has_title"],
        "has_intuition": quality["has_intuition"],
        "has_approach": quality["has_approach"],
        "has_time_complexity": quality["has_time_complexity"],
        "has_space_complexity": quality["has_space_complexity"],
        "has_code_block": quality["has_code_block"],
        "format_score": quality["format_score"],
        "completeness_score": quality["completeness_score"],
        "manual_edit_distance": manual_edit_distance,
        "accepted_for_posting": accepted_for_posting,
        "error_message": str(error_message or ""),
    }


def log_run_record(record: Dict[str, Any], export_excel: bool = True) -> None:
    ensure_metrics_storage()
    conn = _connect()
    try:
        placeholders = ", ".join("?" for _ in RUN_COLUMNS)
        columns = ", ".join(RUN_COLUMNS)
        values = [record.get(col) for col in RUN_COLUMNS]
        conn.execute(
            f"INSERT INTO llm_runs ({columns}) VALUES ({placeholders})",
            values,
        )
        conn.commit()
    finally:
        conn.close()

    if export_excel:
        export_runs_to_excel()


def estimate_edit_distance(original_text: str, edited_text: str) -> int:
    """Approximate edit distance using sequence ratio and max length."""
    original = original_text or ""
    edited = edited_text or ""
    if not original and not edited:
        return 0
    ratio = SequenceMatcher(a=original, b=edited).ratio()
    return int(round((1 - ratio) * max(len(original), len(edited))))


def update_run_feedback(
    run_id: str,
    accepted_for_posting: Optional[int] = None,
    manual_edit_distance: Optional[int] = None,
    export_excel: bool = True,
) -> None:
    ensure_metrics_storage()
    conn = _connect()
    try:
        conn.execute(
            """
            UPDATE llm_runs
            SET accepted_for_posting = COALESCE(?, accepted_for_posting),
                manual_edit_distance = COALESCE(?, manual_edit_distance)
            WHERE run_id = ?
            """,
            (accepted_for_posting, manual_edit_distance, run_id),
        )
        conn.commit()
    finally:
        conn.close()

    if export_excel:
        export_runs_to_excel()


def fetch_recent_runs(limit: int = 200) -> List[Dict[str, Any]]:
    ensure_metrics_storage()
    conn = _connect()
    try:
        cursor = conn.execute(
            f"SELECT {', '.join(RUN_COLUMNS)} FROM llm_runs ORDER BY id DESC LIMIT ?",
            (limit,),
        )
        rows = cursor.fetchall()
        return [dict(row) for row in rows]
    finally:
        conn.close()


def fetch_metrics_summary() -> Dict[str, Any]:
    ensure_metrics_storage()
    conn = _connect()
    try:
        total_runs = conn.execute("SELECT COUNT(*) FROM llm_runs").fetchone()[0]
        success_runs = conn.execute(
            "SELECT COUNT(*) FROM llm_runs WHERE COALESCE(error_type, '') = ''"
        ).fetchone()[0]
        failed_runs = total_runs - success_runs
        timeout_runs = conn.execute(
            "SELECT COUNT(*) FROM llm_runs WHERE timeout_flag = 1"
        ).fetchone()[0]

        avg_tokens_per_sec = conn.execute(
            "SELECT AVG(tokens_per_sec) FROM llm_runs WHERE tokens_per_sec > 0"
        ).fetchone()[0]
        avg_total_duration_ms = conn.execute(
            "SELECT AVG(total_duration_ms) FROM llm_runs WHERE total_duration_ms > 0"
        ).fetchone()[0]
        avg_completeness = conn.execute(
            "SELECT AVG(completeness_score) FROM llm_runs WHERE completeness_score IS NOT NULL"
        ).fetchone()[0]
        avg_format = conn.execute(
            "SELECT AVG(format_score) FROM llm_runs WHERE format_score IS NOT NULL"
        ).fetchone()[0]

        return {
            "total_runs": total_runs,
            "success_runs": success_runs,
            "failed_runs": failed_runs,
            "timeout_runs": timeout_runs,
            "avg_tokens_per_sec": round(avg_tokens_per_sec or 0.0, 2),
            "avg_total_duration_ms": round(avg_total_duration_ms or 0.0, 2),
            "avg_completeness_score": round(avg_completeness or 0.0, 2),
            "avg_format_score": round(avg_format or 0.0, 2),
        }
    finally:
        conn.close()


def export_runs_to_excel() -> Dict[str, str]:
    ensure_metrics_storage()
    paths = get_metrics_paths()

    conn = _connect()
    try:
        cursor = conn.execute(
            f"SELECT {', '.join(RUN_COLUMNS)} FROM llm_runs ORDER BY id ASC"
        )
        rows = [_prepare_export_row(dict(row)) for row in cursor.fetchall()]
    finally:
        conn.close()

    wb = Workbook()
    usage_ws = wb.active
    usage_ws.title = "Usage"
    usage_ws.append(RUN_COLUMNS)

    for row in rows:
        usage_ws.append([row.get(col) for col in RUN_COLUMNS])

    _autofit_columns(usage_ws)

    summary = fetch_metrics_summary()
    summary_ws = wb.create_sheet(title="Summary")
    summary_ws.append(["Metric", "Value"])
    for key, value in summary.items():
        summary_ws.append([key, value])

    prompt_ws = wb.create_sheet(title="PromptVersionSummary")
    prompt_ws.append(
        [
            "prompt_version",
            "prompt_strategy",
            "model",
            "prompt_hash",
            "runs",
            "avg_completeness",
            "avg_format",
            "avg_tokens_per_sec",
            "avg_output_input_ratio",
        ]
    )

    conn = _connect()
    try:
        grouped = conn.execute(
            """
            SELECT
                COALESCE(prompt_version, '') AS prompt_version,
                COALESCE(prompt_strategy, '') AS prompt_strategy,
                COALESCE(model, '') AS model,
                COALESCE(prompt_hash, '') AS prompt_hash,
                COUNT(*) AS runs,
                AVG(completeness_score) AS avg_completeness,
                AVG(format_score) AS avg_format,
                AVG(tokens_per_sec) AS avg_tokens_per_sec,
                AVG(output_input_ratio) AS avg_output_input_ratio
            FROM llm_runs
            GROUP BY
                COALESCE(prompt_version, ''),
                COALESCE(prompt_strategy, ''),
                COALESCE(model, ''),
                COALESCE(prompt_hash, '')
            ORDER BY runs DESC
            """
        ).fetchall()
    finally:
        conn.close()

    for row in grouped:
        prompt_ws.append(
            [
                row["prompt_version"],
                row["prompt_strategy"],
                row["model"],
                row["prompt_hash"],
                row["runs"],
                round(row["avg_completeness"] or 0.0, 2),
                round(row["avg_format"] or 0.0, 2),
                round(row["avg_tokens_per_sec"] or 0.0, 2),
                round(row["avg_output_input_ratio"] or 0.0, 4),
            ]
        )

    model_ws = wb.create_sheet(title="ModelSummary")
    model_ws.append(["model", "runs", "avg_tokens_per_sec", "avg_duration_ms", "avg_completeness"])

    conn = _connect()
    try:
        model_rows = conn.execute(
            """
            SELECT
                COALESCE(model, '') AS model,
                COUNT(*) AS runs,
                AVG(tokens_per_sec) AS avg_tokens_per_sec,
                AVG(total_duration_ms) AS avg_duration_ms,
                AVG(completeness_score) AS avg_completeness
            FROM llm_runs
            GROUP BY COALESCE(model, '')
            ORDER BY runs DESC
            """
        ).fetchall()
    finally:
        conn.close()

    for row in model_rows:
        model_ws.append(
            [
                row["model"],
                row["runs"],
                round(row["avg_tokens_per_sec"] or 0.0, 2),
                round(row["avg_duration_ms"] or 0.0, 2),
                round(row["avg_completeness"] or 0.0, 2),
            ]
        )

    # Legend sheet — explains field meanings and legacy placeholder values
    legend_ws = wb.create_sheet(title="Legend")
    legend_ws.append(["Field / Value", "Meaning"])
    legend_data = [
        ["prompt_strategy = 'legacy'", "Run logged before prompt strategy tracking was added (pre-v2)"],
        ["problem_link = '\u2014'", "Problem link was not captured (pre-v2 run)"],
        ["prompt_text = '\u2014'", "Prompt text was not captured (pre-v2 run)"],
        ["code_text = '\u2014'", "Solution code was not captured (pre-v2 run)"],
        ["llm_response_text = '\u2014'", "Raw LLM response was not captured (pre-v2 run)"],
        ["code_appended_externally = 1", "Code was appended locally (not generated by LLM)"],
        ["llm_returned_code_block = 1", "LLM included a code block despite instructions not to"],
        ["format_score", "Weighted score 0\u2013100 based on presence of required markdown sections"],
        ["completeness_score", "Percentage of required sections present in the output"],
        ["output_input_ratio", "response_tokens / prompt_tokens \u2014 higher = more verbose output"],
        ["tokens_per_sec", "LLM generation speed"],
        ["", ""],
        ["NOTE", "All rows with '\u2014' are legacy runs. Generate a new post to see fully-populated data."],
    ]
    for entry in legend_data:
        legend_ws.append(entry)
    _autofit_columns(legend_ws)

    primary_path = paths["excel_path"]
    try:
        wb.save(primary_path)
        return {"path": primary_path, "status": "ok", "message": ""}
    except PermissionError:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback_path = os.path.join(paths["stats_dir"], f"token_usage_{timestamp}.xlsx")
        try:
            wb.save(fallback_path)
            return {
                "path": fallback_path,
                "status": "permission_error",
                "message": (
                    f"token_usage.xlsx is open in another program. "
                    f"Saved to {os.path.basename(fallback_path)} instead. "
                    "Close the file in Excel and re-export to overwrite the main file."
                ),
            }
        except Exception as exc:
            return {
                "path": "",
                "status": "error",
                "message": f"Export failed: {exc}",
            }
    except Exception as exc:
        return {"path": "", "status": "error", "message": f"Export failed: {exc}"}

