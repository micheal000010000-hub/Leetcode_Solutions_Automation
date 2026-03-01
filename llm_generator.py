from config import GITHUB_REPO_URL
import requests
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

OLLAMA_URL = "http://localhost:11434/api/generate"
OLLAMA_MODEL = "mistral"


def log_token_usage(problem_number, problem_name, response_data):
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    stats_folder = os.path.join(BASE_DIR, "llm_stats")
    os.makedirs(stats_folder, exist_ok=True)

    excel_path = os.path.join(stats_folder, "token_usage.xlsx")

    # Extract metrics safely
    prompt_tokens = response_data.get("prompt_eval_count", 0)
    response_tokens = response_data.get("eval_count", 0)
    total_duration = response_data.get("total_duration", 0)
    load_duration = response_data.get("load_duration", 0)
    prompt_eval_duration = response_data.get("prompt_eval_duration", 0)
    eval_duration = response_data.get("eval_duration", 0)

    total_tokens = prompt_tokens + response_tokens

    # Convert ns → ms
    total_ms = total_duration / 1_000_000
    load_ms = load_duration / 1_000_000
    prompt_ms = prompt_eval_duration / 1_000_000
    eval_ms = eval_duration / 1_000_000

    # Compute throughput
    tokens_per_sec = 0
    if eval_duration > 0:
        tokens_per_sec = response_tokens / (eval_duration / 1_000_000_000)

    # Create file if not exists
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Usage"
        ws.append([
            "Timestamp",
            "Problem Number",
            "Problem Name",
            "Model",
            "Prompt Tokens",
            "Response Tokens",
            "Total Tokens",
            "Total Duration (ms)",
            "Load Duration (ms)",
            "Prompt Eval Duration (ms)",
            "Generation Duration (ms)",
            "Tokens/sec"
        ])
        wb.save(excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active

    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        problem_number,
        problem_name,
        OLLAMA_MODEL,
        prompt_tokens,
        response_tokens,
        total_tokens,
        round(total_ms, 2),
        round(load_ms, 2),
        round(prompt_ms, 2),
        round(eval_ms, 2),
        round(tokens_per_sec, 2)
    ])

    wb.save(excel_path)

def generate_solution_post(problem_number, problem_name, difficulty, link, code, language):
    """
    Generates structured LeetCode solution post using Mistral via Ollama.
    Logs token usage into Excel.
    """

    try:
        prompt = f"""You are an expert technical writer creating a high-quality LeetCode solution post.

    Problem Details:
    Number: {problem_number}
    Name: {problem_name}
    Difficulty: {difficulty}
    Link: {link}

    {language} Solution:
    {code}

    Instructions:

    1. Generate a strong, descriptive, professional Title.
    - The title MUST mention:
        • The core technique used
        • The time complexity (Big-O notation)
    - The code must not be changed
    - The explanation must match the provided language

    2. Generate structured Markdown sections:

    ## Intuition
    ## Approach
    ## Time Complexity
    ## Space Complexity
    ## Code

    3. Format code exactly:

    ```{language}
    {code}

    ````
    """

        response = requests.post(
            OLLAMA_URL,
            json={
                "model": OLLAMA_MODEL,
                "num_predict": 800,
                "prompt": prompt,
                "stream": False,
                "temperature": 0.2,
            },
            timeout=600,
        )

        if response.status_code != 200:
            return f"⚠ Ollama returned status code {response.status_code}"

        response_data = response.json()

        generated_text = response_data.get("response", "").strip()

        # 🔥 Extract token data
        prompt_tokens = response_data.get("prompt_eval_count")
        response_tokens = response_data.get("eval_count")
        total_duration = response_data.get("total_duration")

        # 🔥 Log full response data to Excel
        log_token_usage(
            problem_number,
            problem_name,
            response_data
        )

        if GITHUB_REPO_URL:
            generated_text += (
                "\n\n---\n"
                f"🔗 **GitHub Repository:** {GITHUB_REPO_URL}\n"
            )

        if not generated_text:
            return "⚠ Mistral returned empty response."

        return generated_text

    except requests.exceptions.Timeout:
        return "⚠ Request timed out."
    except requests.exceptions.RequestException as e:
        return f"⚠ Network error: {str(e)}"
    except Exception as e:
        return f"⚠ Error generating solution post: {str(e)}"